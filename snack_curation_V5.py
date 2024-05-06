#23.07.20
#1. 개수 올림처리하여 정수로 표기
#2. 금액의 기준을 가격이 아닌 마진금액(32%) 열에 맞추기
#3. 큐레이션 상품 총 합계금액을 각 시트에 표기
import sys
from PyQt5 import QtCore,QtGui,QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import QSize, Qt
from PyQt5.QtWidgets import *
from PyQt5 import uic
import pandas as pd
import openpyxl
import datetime as dt
import random
import math

#UI파일 연결
#단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
form_class = uic.loadUiType("curation_ui.ui")[0]

#화면을 띄우는데 사용되는 Class 선언
#class WindowClass(QMainWindow, form_class) :
class WindowClass(QDialog) :
    def __init__(self) :
        super().__init__()
        #self.setupUi(self)
        uic.loadUi("curation_ui.ui",self)
        self.setWindowTitle("스낵365 큐레이션")
        self.pushButton_RUN.clicked.connect(self.run_curation)
        self.pushButton_search_file.clicked.connect(self.search_file)

    #큐레이션 실행
    def run_curation(self):

        #원가표 불러오기
        snacklist = pd.read_excel(filename[0],sheet_name='품목리스트 최종',engine='openpyxl',dtype ='str')
        

        #snacklist를 str로 불러오기 때문에 숫자는 int형으로 변환
        selected_columns = ['마진금액 (32%)', '개수']
        snacklist[selected_columns] = snacklist[selected_columns].astype(int)
        
        #원가표의 카테고리 리스트화
        snacklist['카테고리종합'] = ''
        for i in range(len(snacklist)):
            snacklist['카테고리종합'][i] = str(snacklist['메인카테고리'][i]) + '/' + str(snacklist['대(중복선택가능)'][i]) + '/' + str(snacklist['중(중복선택가능)'][i]) + '/' + str(snacklist['소(중복선택가능)'][i]) + '/' + str(snacklist['극소(중복선택가능)'][i])
            snacklist['카테고리종합'][i] = snacklist['카테고리종합'][i].replace('/nan','')
            snacklist['카테고리종합'][i] = snacklist['카테고리종합'][i].split('/')

        #업체정보 받아오기
        wb = openpyxl.load_workbook(filename[0])
        sheetnames = wb.sheetnames
        new_wb = openpyxl.Workbook() #새 엑셀파일 만들기
        snack_합계 = pd.DataFrame(columns=['상품코드','상품명','수량','상온박스','업체명'])
        x = dt.datetime.now()
        new_filename = str(x.year) + '.' + str(x.month) + '.' + str(x.day) + '_' + str(x.hour) + '시' + str(x.minute) + '분' + str(x.second) + '초' + '큐레이션생성.xlsx'

        # 상품명 변경 전에 snacklist 복사본 만들어두기
        snacklist_copy = snacklist.copy()

        # 공급처가 '웰'인 경우 상품명 앞에 '웰'이 붙도록 수정
        for k in range(len(snacklist)):
            if (snacklist['공급처'][k]=='웰' and snacklist['상품명'][k][:2] != '웰 '):
                snacklist['상품명'][k] = '웰 ' + snacklist['상품명'][k]


        for company in sheetnames:
            choice0 = snacklist.copy()
            if company == '카테고리' or company == "품목리스트 최종" or company == "업체정보 최신":
                continue
            sht = wb[company] #시트 선언
            company_name = sht.cell(row=4, column=8).value #업체명

            snack_box = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','수량','총금액'])  # 카테고리에 따른 스낵 후보

            total_price = sht.cell(row=6, column=8).value 

            #230724 포장용기 옵션 선택
            box_option = sht.cell(row=6, column=12).value
            if box_option == '상온박스':
                box_option = '상온박스'
            else:
                box_option = '택배박스'

            #230801 상품용량 선택
            snack_volume = sht.cell(row=5, column=12).value
            if snack_volume == '대':
                choice0 = choice0.drop(choice0[choice0['대/소/전체']=='소'].index)
            elif snack_volume == '소':
                choice0 = choice0.drop(choice0[choice0['대/소/전체']=='대'].index)

            #비선호상품 삭제
            for i in range(50):
                if sht.cell(row=i+10, column=13).value is None:
                    break
                unlike_snack = sht.cell(row=i+10, column=13).value
                choice0 = choice0.drop(choice0[choice0['상품명']==unlike_snack].index)
            for i in range(50):
                if sht.cell(row=i+10, column=14).value is None:
                    break
                unlike_snack = sht.cell(row=i+10, column=14).value
                choice0 = choice0.drop(choice0[choice0['상품명']==unlike_snack].index)

            #필수상품 우선 추가
            for i in range(100):
                if sht.cell(row=i+11, column=11).value is None:
                    break
                essential_snack = sht.cell(row=i+11, column=11).value
                if (snacklist_copy.loc[snacklist_copy['상품명']==essential_snack]['공급처'].item()=='웰'):  #필수상품의 공급처가 '웰'인 경우에 상품명 앞에 '웰'을 붙여야 함
                    essential_snack = '웰 ' + essential_snack
                price = choice0.loc[choice0['상품명']==essential_snack]['마진금액 (32%)'].item()
                quantity = math.ceil(sht.cell(row=i+11, column=12).value) #개수
                choice0 = choice0.drop(choice0[choice0['상품명']==essential_snack].index)
                if total_price >= price*quantity:
                    total_price -= price * quantity
                    if essential_snack in snack_box['상품명'].values:
                        snack_box.loc[snack_box['상품명']==essential_snack,'수량'] += quantity
                        snack_box.loc[snack_box['상품명']==essential_snack,'총금액'] += price * quantity
                    else:
                        snack_box = snack_box.append({'상품코드':str(snacklist.loc[snacklist['상품명']==essential_snack]['상품코드'].item()),'상품명':essential_snack, '마진금액 (32%)': price, '수량':quantity, '총금액':price * quantity}, ignore_index=True)
                else:
                    break
            
            choice0 = choice0.reset_index(drop=True, inplace=False)
            
                    


            for i in range(10,21):                       #업체 시트의 카테고리 종류 탐색
                customer_kategorie_메인 = pd.DataFrame(columns=['A'])
                customer_kategorie_대 = pd.DataFrame(columns=['A'])
                customer_kategorie_중 = pd.DataFrame(columns=['A'])
                customer_kategorie_소 = pd.DataFrame(columns=['A'])
                customer_kategorie_극소 = pd.DataFrame(columns=['A'])
                choice1 = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])
                choice2 = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])
                choice3 = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])
                choice4 = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])
                choice5 = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])

                #카테고리 단계별로 데이터프레임화
                temp_kategorie=wb[company].cell(row=i, column=3).value  # 메인카테고리
                if temp_kategorie is None:
                    customer_kategorie_메인 = customer_kategorie_메인.append({'A': ''}, ignore_index=True)
                else:
                    if '/' in temp_kategorie:
                        customer_kategorie_메인 = customer_kategorie_메인.append({'A': temp_kategorie.split('/')}, ignore_index=True)
                    else:
                        customer_kategorie_메인 = customer_kategorie_메인.append({'A': temp_kategorie}, ignore_index=True)

                temp_kategorie=wb[company].cell(row=i, column=4).value  # 대 카테고리
                if temp_kategorie is None:
                    customer_kategorie_대 = customer_kategorie_대.append({'A': ''}, ignore_index=True)
                else:
                    if '/' in temp_kategorie:
                        customer_kategorie_대 = customer_kategorie_대.append({'A': temp_kategorie.split('/')}, ignore_index=True)
                    else:
                        customer_kategorie_대 = customer_kategorie_대.append({'A': temp_kategorie}, ignore_index=True)

                temp_kategorie=wb[company].cell(row=i, column=5).value  # 중 카테고리
                if temp_kategorie is None:
                    customer_kategorie_중 = customer_kategorie_중.append({'A': ''}, ignore_index=True)
                else:
                    if '/' in temp_kategorie:
                        customer_kategorie_중 = customer_kategorie_중.append({'A': temp_kategorie.split('/')}, ignore_index=True)
                    else:
                        customer_kategorie_중 = customer_kategorie_중.append({'A': temp_kategorie}, ignore_index=True)
                
                temp_kategorie=wb[company].cell(row=i, column=6).value  # 소 카테고리
                if temp_kategorie is None:
                    customer_kategorie_소 = customer_kategorie_소.append({'A': ''}, ignore_index=True)
                else:
                    if '/' in temp_kategorie:
                        customer_kategorie_소 = customer_kategorie_소.append({'A': temp_kategorie.split('/')}, ignore_index=True)
                    else:
                        customer_kategorie_소 = customer_kategorie_소.append({'A': temp_kategorie}, ignore_index=True)

                temp_kategorie=wb[company].cell(row=i, column=7).value  # 극소 카테고리
                if temp_kategorie is None:
                    customer_kategorie_극소 = customer_kategorie_극소.append({'A': ''}, ignore_index=True)
                else:
                    if '/' in temp_kategorie:
                        customer_kategorie_극소 = customer_kategorie_극소.append({'A': temp_kategorie.split('/')}, ignore_index=True)
                    else:
                        customer_kategorie_극소 = customer_kategorie_극소.append({'A': temp_kategorie}, ignore_index=True)

                customer_kategorie_비율 = wb[company].cell(row=i, column=8).value


                #snacklist를 탐색해서 큰 카테고리에 해당하는 상품 이동 (아래 class로 상품 복사가 아님)
                for k in range(len(choice0)):
                    
                    #print(type(customer_kategorie_메인['A']))
                    if isinstance(customer_kategorie_메인['A'].iloc[0], list):
                        elements_kategorie = set(customer_kategorie_메인['A'].iloc[0])
                    else:
                        elements_kategorie = set([customer_kategorie_메인['A'].iloc[0]])
                    #elements_kategorie = set(tuple(row) for row in customer_kategorie_메인['A'])
                    #elements_kategorie = set(list(customer_kategorie_메인['A']))
                    #if isinstance(customer_kategorie_메인['A'], list):
                    #    elements_kategorie = set(customer_kategorie_메인['A'].item())
                    #else:
                    #    elements_kategorie = set(customer_kategorie_메인['A'])
                    elements_snack = set(choice0['카테고리종합'][k])
                    if elements_kategorie.intersection(elements_snack):
                        if math.ceil(choice0['개수'][k]) == 0: #개수가 0개인 상품 삭제
                            continue
                        choice1 = choice1.append({'상품코드':choice0['상품코드'][k], '상품명':choice0['상품명'][k], '마진금액 (32%)':choice0['마진금액 (32%)'][k], '개수':math.ceil(choice0['개수'][k]), '카테고리종합':choice0['카테고리종합'][k]}, ignore_index=True)
                #choice1을 탐색해서 카테고리 '대'에 해당하는 상품 추출('대'가 비어있으면 choice2=choice1)
                choice_temp = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])
                for m in range(len(choice1)):
                    if isinstance(customer_kategorie_대['A'].iloc[0], list):
                        elements_kategorie = set(customer_kategorie_대['A'].iloc[0])
                    else:
                        elements_kategorie = set([customer_kategorie_대['A'].iloc[0]])
                    #elements_kategorie = set(list(customer_kategorie_대['A']))
                    #if isinstance(customer_kategorie_대['A'], list):
                    #    elements_kategorie = set(customer_kategorie_대['A'])
                    #else:
                    #    elements_kategorie = set([customer_kategorie_대['A']])
                    elements_snack = choice1['카테고리종합'][m]
                    if '' in elements_kategorie:    # '대'카테고리가 선정되어있지 않으면 상위 카테고리 해당 물품 전체 가져오기
                        choice2 = choice1
                        break
                    if elements_kategorie.intersection(elements_snack):
                        choice2 = choice2.append({'상품코드':choice1['상품코드'][m], '상품명':choice1['상품명'][m], '마진금액 (32%)':choice1['마진금액 (32%)'][m], '개수':choice1['개수'][m], '카테고리종합':choice1['카테고리종합'][m]}, ignore_index=True)
                    else:
                        choice_temp = choice_temp.append({'상품코드':choice1['상품코드'][m], '상품명':choice1['상품명'][m], '마진금액 (32%)':choice1['마진금액 (32%)'][m], '개수':choice1['개수'][m], '카테고리종합':choice1['카테고리종합'][m]}, ignore_index=True)
                choice1 = choice_temp   #choice1에서 choice2로 분류된 상품은 제거
                #choice2를 탐색해서 카테고리 '중'에 해당하는 상품 추출('중'이 비어있으면 choice3=choice2)
                choice_temp = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])
                for n in range(len(choice2)):
                    if isinstance(customer_kategorie_중['A'].iloc[0], list):
                        elements_kategorie = set(customer_kategorie_중['A'].iloc[0])
                    else:
                        elements_kategorie = set([customer_kategorie_중['A'].iloc[0]])
                    #elements_kategorie = set(customer_kategorie_중['A'])
                    #if isinstance(customer_kategorie_중['A'], list):
                    #    elements_kategorie = set(customer_kategorie_중['A'])
                    #else:
                    #    elements_kategorie = set([customer_kategorie_중['A']])
                    elements_snack = choice2['카테고리종합'][n]
                    if '' in elements_kategorie:    # '중'카테고리가 선정되어있지 않으면 상위 카테고리 해당 물품 전체 가져오기
                        choice3 = choice2
                        break
                    if elements_kategorie.intersection(elements_snack):
                        choice3 = choice3.append({'상품코드':choice2['상품코드'][n], '상품명':choice2['상품명'][n], '마진금액 (32%)':choice2['마진금액 (32%)'][n], '개수':choice2['개수'][n], '카테고리종합':choice2['카테고리종합'][n]}, ignore_index=True)
                    else:
                        choice_temp = choice_temp.append({'상품코드':choice2['상품코드'][n], '상품명':choice2['상품명'][n], '마진금액 (32%)':choice2['마진금액 (32%)'][n], '개수':choice2['개수'][n], '카테고리종합':choice2['카테고리종합'][n]}, ignore_index=True)
                choice2 = choice_temp   #choice2에서 choice3로 분류된 상품은 제거
                #choice3을 탐색해서 카테고리 '소'에 해당하는 상품 추출('소'가 비어있으면 choice4=choice3)
                choice_temp = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])
                for p in range(len(choice3)):
                    if isinstance(customer_kategorie_소['A'].iloc[0], list):
                        elements_kategorie = set(customer_kategorie_소['A'].iloc[0])
                    else:
                        elements_kategorie = set([customer_kategorie_소['A'].iloc[0]])
                    #elements_kategorie = set(customer_kategorie_소['A'])
                    #if isinstance(customer_kategorie_소['A'], list):
                    #    elements_kategorie = set(customer_kategorie_소['A'])
                    #else:
                    #    elements_kategorie = set([customer_kategorie_소['A']])
                    elements_snack = choice3['카테고리종합'][p]
                    if '' in elements_kategorie:    # '소'카테고리가 선정되어있지 않으면 상위 카테고리 해당 물품 전체 가져오기
                        choice4 = choice3
                        break
                    if elements_kategorie.intersection(elements_snack):
                        choice4 = choice4.append({'상품코드':choice3['상품코드'][p], '상품명':choice3['상품명'][p], '마진금액 (32%)':choice3['마진금액 (32%)'][p], '개수':choice3['개수'][p], '카테고리종합':choice3['카테고리종합'][p]}, ignore_index=True)
                    else:
                        choice_temp = choice_temp.append({'상품코드':choice3['상품코드'][p], '상품명':choice3['상품명'][p], '마진금액 (32%)':choice3['마진금액 (32%)'][p], '개수':choice3['개수'][p], '카테고리종합':choice3['카테고리종합'][p]}, ignore_index=True)
                choice3 = choice_temp   #choice3에서 choice4로 분류된 상품은 제거
                #choice4를 탐색해서 카테고리 '극소'에 해당하는 상품 추출('극소'가 비어있으면 choice5=choice4)
                choice_temp = pd.DataFrame(columns=['상품코드','상품명','마진금액 (32%)','개수','카테고리종합'])
                for q in range(len(choice4)):
                    if isinstance(customer_kategorie_극소['A'].iloc[0], list):
                        elements_kategorie = set(customer_kategorie_극소['A'].iloc[0])
                    else:
                        elements_kategorie = set([customer_kategorie_극소['A'].iloc[0]])
                    #elements_kategorie = set(customer_kategorie_극소['A'])
                    #if isinstance(customer_kategorie_극소['A'], list):
                    #    elements_kategorie = set(customer_kategorie_극소['A'])
                    #else:
                    #    elements_kategorie = set([customer_kategorie_극소['A']])
                    elements_snack = choice4['카테고리종합'][q]
                    if '' in elements_kategorie:    # '극소'카테고리가 선정되어있지 않으면 상위 카테고리 해당 물품 전체 가져오기
                        choice5 = choice4
                        break
                    if elements_kategorie.intersection(elements_snack):
                        choice5 = choice5.append({'상품코드':choice4['상품코드'][q], '상품명':choice4['상품명'][q], '마진금액 (32%)':choice4['마진금액 (32%)'][q], '개수':choice4['개수'][q], '카테고리종합':choice4['카테고리종합'][q]}, ignore_index=True)
                    else:
                        choice_temp = choice_temp.append({'상품코드':choice4['상품코드'][q], '상품명':choice4['상품명'][q], '마진금액 (32%)':choice4['마진금액 (32%)'][q], '개수':choice4['개수'][q], '카테고리종합':choice4['카테고리종합'][q]}, ignore_index=True)
                choice4 = choice_temp   #choice4에서 choice5로 분류된 상품은 제거


                #퍼센트에 따라서 choice5에 선정된 후보 상품으로 구성해야할 금액
                if customer_kategorie_비율 is None:
                    continue
                kategorie_price = total_price * customer_kategorie_비율 / 100
                
                #kategorie_price에 해당하는 금액에 가깝게 상품 구성, choice레벨마다 상품이 없으면 상위 상품으로 이동
                while kategorie_price > 0:
                    if len(choice1)+len(choice2)+len(choice3)+len(choice4)+len(choice5) == 0:
                        break
                    while len(choice5)>0:
                        choice5 = choice5.reset_index(drop=True, inplace=False)
                        random_snack = random.choice(choice5['상품명'])
                        price = choice5.loc[choice5['상품명']==random_snack]['마진금액 (32%)'].item()
                        quantity = choice5.loc[choice5['상품명']==random_snack]['개수'].item() #개수
                        choice5 = choice5.drop(choice5[choice5['상품명']==random_snack].index)
                        if kategorie_price >= price*quantity:
                            kategorie_price -= price * quantity
                            if random_snack in snack_box['상품명'].values:
                                snack_box.loc[snack_box['상품명']==random_snack,'수량'] += quantity
                                snack_box.loc[snack_box['상품명']==random_snack,'총금액'] += price * quantity
                            else:
                                snack_box = snack_box.append({'상품코드':snacklist.loc[snacklist['상품명']==random_snack]['상품코드'].item(),'상품명':random_snack, '마진금액 (32%)': price, '수량':quantity, '총금액':price * quantity}, ignore_index=True)
                        else:
                            break
                    while len(choice4)>0:
                        choice4 = choice4.reset_index(drop=True, inplace=False)
                        random_snack = random.choice(choice4['상품명'])
                        price = choice4.loc[choice4['상품명']==random_snack]['마진금액 (32%)'].item()
                        quantity = choice4.loc[choice4['상품명']==random_snack]['개수'].item() #개수
                        choice4 = choice4.drop(choice4[choice4['상품명']==random_snack].index)
                        if kategorie_price >= price*quantity:
                            kategorie_price -= price * quantity
                            if random_snack in snack_box['상품명'].values:
                                snack_box.loc[snack_box['상품명']==random_snack,'수량'] += quantity
                                snack_box.loc[snack_box['상품명']==random_snack,'총금액'] += price * quantity
                            else:
                                snack_box = snack_box.append({'상품코드':snacklist.loc[snacklist['상품명']==random_snack]['상품코드'].item(),'상품명':random_snack, '마진금액 (32%)': price, '수량':quantity, '총금액':price * quantity}, ignore_index=True)
                        else:
                            break
                    while len(choice3)>0:
                        choice3 = choice3.reset_index(drop=True, inplace=False)
                        random_snack = random.choice(choice3['상품명'])
                        price = choice3.loc[choice3['상품명']==random_snack]['마진금액 (32%)'].item()
                        quantity = choice3.loc[choice3['상품명']==random_snack]['개수'].item() #개수
                        choice3 = choice3.drop(choice3[choice3['상품명']==random_snack].index)
                        if kategorie_price >= price*quantity:
                            kategorie_price -= price * quantity
                            if random_snack in snack_box['상품명'].values:
                                snack_box.loc[snack_box['상품명']==random_snack,'수량'] += quantity
                                snack_box.loc[snack_box['상품명']==random_snack,'총금액'] += price * quantity
                            else:
                                snack_box = snack_box.append({'상품코드':snacklist.loc[snacklist['상품명']==random_snack]['상품코드'].item(),'상품명':random_snack, '마진금액 (32%)': price, '수량':quantity, '총금액':price * quantity}, ignore_index=True)
                        else:
                            break
                    while len(choice2)>0:
                        choice2 = choice2.reset_index(drop=True, inplace=False)
                        random_snack = random.choice(choice2['상품명'])
                        price = choice2.loc[choice2['상품명']==random_snack]['마진금액 (32%)'].item()
                        quantity = choice2.loc[choice2['상품명']==random_snack]['개수'].item() #개수
                        choice2 = choice2.drop(choice2[choice2['상품명']==random_snack].index)
                        if kategorie_price >= price*quantity:
                            kategorie_price -= price * quantity
                            if random_snack in snack_box['상품명'].values:
                                snack_box.loc[snack_box['상품명']==random_snack,'수량'] += quantity
                                snack_box.loc[snack_box['상품명']==random_snack,'총금액'] += price * quantity
                            else:
                                snack_box = snack_box.append({'상품코드':snacklist.loc[snacklist['상품명']==random_snack]['상품코드'].item(),'상품명':random_snack, '마진금액 (32%)': price, '수량':quantity, '총금액':price * quantity}, ignore_index=True)
                        else:
                            break
                    while len(choice1)>0:
                        choice1 = choice1.reset_index(drop=True, inplace=False)
                        random_snack = random.choice(choice1['상품명'])
                        price = choice1.loc[choice1['상품명']==random_snack]['마진금액 (32%)'].item()
                        quantity = choice1.loc[choice1['상품명']==random_snack]['개수'].item() #개수
                        choice1 = choice1.drop(choice1[choice1['상품명']==random_snack].index)
                        if kategorie_price >= price*quantity:
                            kategorie_price -= price * quantity
                            if random_snack in snack_box['상품명'].values:
                                snack_box.loc[snack_box['상품명']==random_snack,'수량'] += quantity
                                snack_box.loc[snack_box['상품명']==random_snack,'총금액'] += price * quantity
                            else:
                                snack_box = snack_box.append({'상품코드':snacklist.loc[snacklist['상품명']==random_snack]['상품코드'].item(),'상품명':random_snack, '마진금액 (32%)': price, '수량':quantity, '총금액':price * quantity}, ignore_index=True)
                        else:
                            break
                            
            #상품명에 따른 오름차순 정렬
            snack_box = snack_box.sort_values(by='상품명') 
            snack_box = snack_box.reset_index(drop=True)
            
            #새로운 파일에 시트 생성
            new_wb.create_sheet(company)
            new_wb[company].cell(row=1, column=1).value = '업체명'
            new_wb[company].cell(row=1, column=2).value = company_name
            new_wb[company].column_dimensions['B'].width = 50
            new_wb[company].cell(row=2, column=1).value = '상품코드'
            new_wb[company].cell(row=2, column=2).value = '상품명'
            new_wb[company].cell(row=2, column=3).value = '단품가격'
            new_wb[company].cell(row=2, column=4).value = '수량'
            new_wb[company].cell(row=2, column=5).value = '총금액'
            new_wb[company].cell(row=2, column=6).value = '포장용기'
            new_wb[company].cell(row=2, column=7).value = '업체명'
            new_wb[company].cell(row=2, column=9).value = '합산금액'

            #고객 별 합산금액 변수
            cost_sum = 0
            for i in range(len(snack_box)):
                new_wb[company].cell(row=3+i, column=1).value = snack_box['상품코드'][i]
                new_wb[company].cell(row=3+i, column=2).value = snack_box['상품명'][i]
                new_wb[company].cell(row=3+i, column=3).value = snack_box['마진금액 (32%)'][i]
                new_wb[company].cell(row=3+i, column=4).value = snack_box['수량'][i]
                new_wb[company].cell(row=3+i, column=5).value = snack_box['총금액'][i]
                new_wb[company].cell(row=3+i, column=6).value = box_option
                new_wb[company].cell(row=3+i, column=7).value = company_name
                #if(box_option == '상온박스'):
                #    new_wb[company].cell(row=3+i, column=5).value = '상온박스'
                #    new_wb[company].cell(row=3+i, column=6).value = company_name
                cost_sum = cost_sum + snack_box['총금액'][i]
                if snack_box['상품명'][i] in snack_합계['상품명'].values:
                    snack_합계.loc[snack_합계['상품명']==snack_box['상품명'][i],'수량'] += snack_box['수량'][i]
                else:
                    snack_합계 = snack_합계.append({'상품코드':snack_box['상품코드'][i],'상품명':snack_box['상품명'][i], '수량':snack_box['수량'][i]}, ignore_index=True)
                
                new_wb[company].cell(row=2, column=10).value = cost_sum #합산금액
                
        # 큐레이션 스낵 별 수량 합계        
        new_wb.create_sheet('합계')
        new_wb['합계'].cell(row=1, column=1).value = '상품코드'
        new_wb['합계'].cell(row=1, column=2).value = '상품명'
        new_wb['합계'].cell(row=1, column=3).value = '수량'
        new_wb['합계'].column_dimensions['B'].width = 50
        for i in range(len(snack_합계)):
            new_wb['합계'].cell(row=2+i, column=1).value = snack_합계['상품코드'][i]
            new_wb['합계'].cell(row=2+i, column=2).value = snack_합계['상품명'][i]
            new_wb['합계'].cell(row=2+i, column=3).value = snack_합계['수량'][i]

        del new_wb['Sheet']
        new_wb.save(new_filename) # 새 엑셀파일 이름 변경


    #큐레이션 정보파일 불러오기
    def search_file(self):
        global filename
        filename = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_filename.setText(filename[0])




if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 

    #WindowClass의 인스턴스 생성
    myWindow = WindowClass() 

    #프로그램 화면을 보여주는 코드
    #myWindow.show()
    myWindow.show()

    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    app.exec_()